"""One-time repair: re-derive Startup Library / Network Library fields from the source workbook.

Earlier imports of ``Ekosistem_Library_New.xlsx`` ran with column mappings written for an
older workbook layout (Turkish headers), so fields such as Category, Vertical, Geography,
Source, Status, and Borusan fits were lost or wrong on existing organization records.
This script reads the workbook again with the corrected header mapping and fixes the
records in place.

Safety model:
- DRY RUN by default; pass --apply to write changes.
- A field is overwritten only when the workbook has a value AND either the DB field is
  empty or the organization has never been manually edited (updated_by_user_id is NULL).
  Manually edited records with differing values are reported as conflicts, not changed.
- No organizations are created or deleted. Unmatched workbook rows are reported.
- Every change (and every skipped conflict) is written to a JSON report for review.

Usage (from the backend directory, venv active, DATABASE_URL set):
    python scripts/repair_library_from_workbook.py "C:\\path\\to\\Ekosistem_Library_New.xlsx"
    python scripts/repair_library_from_workbook.py "C:\\path\\to\\Ekosistem_Library_New.xlsx" --apply
"""

import argparse
import json
import sys
from datetime import date, datetime, timezone
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from sqlalchemy import select

BACKEND_ROOT = Path(__file__).resolve().parents[1]
if str(BACKEND_ROOT) not in sys.path:
    sys.path.insert(0, str(BACKEND_ROOT))

from app.db.session import SessionLocal  # noqa: E402
from app.models import (  # noqa: E402
    AuditLog,
    BorusanCompany,
    Organization,
    OrganizationBorusanFit,
    OrganizationTag,
    Tag,
)
from app.services.excel_import.candidates import ImportCandidateService  # noqa: E402

CONFIG_DIR = Path(__file__).resolve().parents[2] / "config"

STARTUP_SHEET = "Startup Library"
NETWORK_SHEET = "Network"

# Startup Library headers (new workbook first, legacy fallback second).
STARTUP_FIELD_HEADERS: dict[str, tuple[str, ...]] = {
    "category": ("Category",),
    "vertical_text": ("Vertical",),
    "geography_text": ("Geography", "Coğrafya"),
    "source_text": ("Source", "Kaynak"),
    "added_by_text": ("Added By",),
    "solution_summary": ("Solution / Use-case",),
    "last_contact_date": ("Last Contact",),
    "lifecycle_status": ("Process Status (1 to 7)", "Status"),
}

NETWORK_FIELD_HEADERS: dict[str, tuple[str, ...]] = {
    "organization_subtype": ("Type",),
    "geography_text": ("Geography",),
    "added_by_text": ("Added By / Contact Owner", "Ekleyen / Görüşen"),
    "relationship_status": ("Relationship",),
    "expertise": ("Expertise",),
}

# Column length limits from the organizations table. Values that do not fit are
# almost always evidence of a shifted row in the workbook itself, so they are
# reported and skipped rather than truncated.
FIELD_MAX_LENGTHS: dict[str, int] = {
    "organization_subtype": 80,
    "category_code": 120,
    "category_label": 255,
    "vertical_text": 255,
    "geography_text": 255,
    "source_text": 512,
    "added_by_text": 255,
}


def clean_cell(value: Any) -> Any:
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, str):
        cleaned = " ".join(value.replace("\xa0", " ").strip().split())
        if cleaned.lower() in {"", "-", "?", "n/a", "na", "none", "null"}:
            return None
        return cleaned
    return value


def sheet_rows(workbook_path: Path, sheet_name: str) -> list[dict[str, Any]]:
    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    try:
        if sheet_name not in workbook.sheetnames:
            return []
        worksheet = workbook[sheet_name]
        rows_iter = worksheet.iter_rows(values_only=True)
        try:
            header_row = next(rows_iter)
        except StopIteration:
            return []
        headers = [str(h) if h is not None else f"__EMPTY_{i}" for i, h in enumerate(header_row)]
        results: list[dict[str, Any]] = []
        for excel_index, row in enumerate(rows_iter, start=2):
            values: dict[str, Any] = {}
            for index, header in enumerate(headers):
                if header.startswith("__EMPTY_"):
                    continue
                cell = clean_cell(row[index]) if index < len(row) else None
                # Keep the first non-empty value for duplicated header names.
                if header in values and values[header] is not None:
                    continue
                values[header] = cell
            if any(value is not None for value in values.values()):
                values["__excel_row"] = excel_index
                results.append(values)
        return results
    finally:
        workbook.close()


def first_value(values: dict[str, Any], names: tuple[str, ...]) -> Any:
    for name in names:
        value = values.get(name)
        if value not in (None, ""):
            return value
    return None


class LibraryRepair:
    def __init__(self, workbook_path: Path, apply_changes: bool) -> None:
        self.workbook_path = workbook_path
        self.apply_changes = apply_changes
        self.service = ImportCandidateService(CONFIG_DIR)
        self.changes: list[dict[str, Any]] = []
        self.conflicts: list[dict[str, Any]] = []
        self.unmatched: list[dict[str, Any]] = []
        self.ambiguous: list[dict[str, Any]] = []
        self.data_quality: list[dict[str, Any]] = []
        self.fit_additions = 0
        self.tag_links_added = 0
        self.orgs_touched: set[str] = set()

    # ------------------------------------------------------------------ helpers

    def _record_change(self, org: Organization, field: str, before: Any, after: Any) -> None:
        self.changes.append(
            {
                "organization_id": str(org.id),
                "name": org.name,
                "field": field,
                "before": self._serialize(before),
                "after": self._serialize(after),
            }
        )
        self.orgs_touched.add(str(org.id))

    def _record_conflict(self, org: Organization, field: str, db_value: Any, workbook_value: Any) -> None:
        self.conflicts.append(
            {
                "organization_id": str(org.id),
                "name": org.name,
                "field": field,
                "db_value": self._serialize(db_value),
                "workbook_value": self._serialize(workbook_value),
                "reason": "Organization was manually edited; not overwritten.",
            }
        )

    @staticmethod
    def _serialize(value: Any) -> Any:
        if isinstance(value, (date, datetime)):
            return value.isoformat()
        return value

    def _record_data_quality(self, org: Organization, field: str, value: Any, reason: str) -> None:
        self.data_quality.append(
            {
                "organization_id": str(org.id),
                "name": org.name,
                "field": field,
                "workbook_value": str(value)[:200],
                "reason": reason,
            }
        )

    def _set_field(self, org: Organization, field: str, new_value: Any) -> None:
        """Apply the overwrite policy for one scalar field."""
        if new_value in (None, ""):
            return
        max_length = FIELD_MAX_LENGTHS.get(field)
        if max_length and isinstance(new_value, str) and len(new_value) > max_length:
            self._record_data_quality(
                org,
                field,
                new_value,
                f"Value exceeds {max_length} characters; the workbook row is probably shifted. Skipped.",
            )
            return
        if field == "geography_text" and isinstance(new_value, str) and ("@" in new_value or len(new_value) > 80):
            self._record_data_quality(
                org,
                field,
                new_value,
                "Geography value looks like contact/free text; the workbook row is probably shifted. Skipped.",
            )
            return
        current = getattr(org, field)
        if isinstance(current, str) and not current.strip():
            current = None
        if current == new_value:
            return
        if current is not None and org.updated_by_user_id is not None:
            self._record_conflict(org, field, current, new_value)
            return
        self._record_change(org, field, current, new_value)
        if self.apply_changes:
            setattr(org, field, new_value)

    def _ensure_tag_link(self, db, org: Organization, label: str, tag_group: str) -> None:
        if not label:
            return
        if self.apply_changes:
            tag = self.service._get_or_create_tag(db, label, tag_group)
            existing = db.execute(
                select(OrganizationTag).where(
                    OrganizationTag.organization_id == org.id,
                    OrganizationTag.tag_id == tag.id,
                )
            ).scalar_one_or_none()
            if existing:
                return
            db.add(OrganizationTag(organization_id=org.id, tag_id=tag.id, source="IMPORT"))
            self.tag_links_added += 1
            self.orgs_touched.add(str(org.id))
        else:
            code = f"{tag_group}_{self.service._slug(label)}"[:120]
            tag = db.execute(select(Tag).where(Tag.code == code)).scalar_one_or_none()
            existing = None
            if tag is not None:
                existing = db.execute(
                    select(OrganizationTag).where(
                        OrganizationTag.organization_id == org.id,
                        OrganizationTag.tag_id == tag.id,
                    )
                ).scalar_one_or_none()
            if existing is None:
                self.tag_links_added += 1
                self.orgs_touched.add(str(org.id))

    def _ensure_fit(self, db, org: Organization, company_by_code: dict[str, BorusanCompany], code: str, raw_value: str) -> None:
        company = company_by_code.get(code)
        if company is None:
            return
        existing = db.execute(
            select(OrganizationBorusanFit).where(
                OrganizationBorusanFit.organization_id == org.id,
                OrganizationBorusanFit.borusan_company_id == company.id,
            )
        ).scalar_one_or_none()
        if existing:
            return
        self.fit_additions += 1
        self.orgs_touched.add(str(org.id))
        self._record_change(org, f"borusan_fit.{code}", None, "RELEVANT")
        if self.apply_changes:
            db.add(
                OrganizationBorusanFit(
                    organization_id=org.id,
                    borusan_company_id=company.id,
                    fit_level="RELEVANT",
                    source="IMPORT",
                    raw_value=raw_value,
                )
            )

    def _match_organization(
        self,
        row: dict[str, Any],
        name_header: str,
        orgs_by_name: dict[str, list[Organization]],
        orgs_by_domain: dict[str, list[Organization]],
        sheet: str,
    ) -> Organization | None:
        name = row.get(name_header)
        normalized = self.service._normalize_name(name)
        if not normalized:
            return None
        domain = self.service._extract_domain(row.get("Website")) if sheet == STARTUP_SHEET else None
        if domain:
            matches = orgs_by_domain.get(domain, [])
            if len(matches) == 1:
                return matches[0]
            if len(matches) > 1:
                self.ambiguous.append({"sheet": sheet, "excel_row": row.get("__excel_row"), "name": name, "reason": f"{len(matches)} organizations share domain {domain}"})
                return None
        matches = orgs_by_name.get(normalized, [])
        if len(matches) == 1:
            return matches[0]
        if len(matches) > 1:
            self.ambiguous.append({"sheet": sheet, "excel_row": row.get("__excel_row"), "name": name, "reason": f"{len(matches)} organizations share normalized name"})
            return None
        self.unmatched.append({"sheet": sheet, "excel_row": row.get("__excel_row"), "name": name})
        return None

    # ------------------------------------------------------------------ main passes

    def repair(self) -> dict[str, Any]:
        startup_rows = sheet_rows(self.workbook_path, STARTUP_SHEET)
        network_rows = sheet_rows(self.workbook_path, NETWORK_SHEET)

        with SessionLocal() as db:
            organizations = db.execute(select(Organization)).scalars().all()
            startups_by_name: dict[str, list[Organization]] = {}
            startups_by_domain: dict[str, list[Organization]] = {}
            network_by_name: dict[str, list[Organization]] = {}
            for org in organizations:
                bucket = network_by_name if org.organization_type == "NETWORK_INSTITUTION" else startups_by_name
                bucket.setdefault(org.normalized_name, []).append(org)
                if org.organization_type != "NETWORK_INSTITUTION" and org.website_domain:
                    startups_by_domain.setdefault(org.website_domain.lower(), []).append(org)

            company_by_code = {
                company.code: company
                for company in db.execute(select(BorusanCompany)).scalars().all()
            }
            fit_columns = self.service._borusan_fit_columns()

            seen_names: set[str] = set()
            processed_org_ids: set[str] = set()
            for row in startup_rows:
                normalized = self.service._normalize_name(row.get("Startup"))
                if not normalized:
                    continue
                if normalized in seen_names:
                    continue  # keep the first occurrence; duplicates were review items at import time
                seen_names.add(normalized)
                org = self._match_organization(row, "Startup", startups_by_name, startups_by_domain, STARTUP_SHEET)
                if org is None:
                    continue
                if str(org.id) in processed_org_ids:
                    # Two workbook rows (e.g. same domain, different names) resolve to
                    # the same organization; keep the first row, report the rest.
                    self.ambiguous.append(
                        {
                            "sheet": STARTUP_SHEET,
                            "excel_row": row.get("__excel_row"),
                            "name": row.get("Startup"),
                            "reason": f"Row resolves to organization '{org.name}' that was already repaired by an earlier row.",
                        }
                    )
                    continue
                processed_org_ids.add(str(org.id))
                self._repair_startup(db, org, row, company_by_code, fit_columns)

            seen_names.clear()
            processed_org_ids.clear()
            for row in network_rows:
                normalized = self.service._normalize_name(row.get("Institution"))
                if not normalized:
                    continue
                if normalized in seen_names:
                    continue
                seen_names.add(normalized)
                org = self._match_organization(row, "Institution", network_by_name, {}, NETWORK_SHEET)
                if org is None:
                    continue
                if str(org.id) in processed_org_ids:
                    self.ambiguous.append(
                        {
                            "sheet": NETWORK_SHEET,
                            "excel_row": row.get("__excel_row"),
                            "name": row.get("Institution"),
                            "reason": f"Row resolves to organization '{org.name}' that was already repaired by an earlier row.",
                        }
                    )
                    continue
                processed_org_ids.add(str(org.id))
                self._repair_network(db, org, row)

            if self.apply_changes:
                db.add(
                    AuditLog(
                        actor_user_id=None,
                        action="DATA_REPAIR",
                        entity_type="ORGANIZATION",
                        entity_id=None,
                        after_data={
                            "script": "repair_library_from_workbook.py",
                            "workbook": self.workbook_path.name,
                            "field_changes": len(self.changes),
                            "organizations_touched": len(self.orgs_touched),
                            "conflicts_skipped": len(self.conflicts),
                        },
                        created_at=datetime.now(timezone.utc),
                    )
                )
                db.commit()
            else:
                db.rollback()

        return self._report()

    def _repair_startup(
        self,
        db,
        org: Organization,
        row: dict[str, Any],
        company_by_code: dict[str, BorusanCompany],
        fit_columns: dict[str, str],
    ) -> None:
        category_value = first_value(row, STARTUP_FIELD_HEADERS["category"])
        category_tags = self.service._split_tags(category_value, "CATEGORY")
        if category_tags:
            self._set_field(org, "category_code", self.service._category_code(category_tags))
            self._set_field(org, "category_label", category_tags[0]["label"])
            for tag_payload in category_tags:
                self._ensure_tag_link(db, org, tag_payload["label"], "CATEGORY")

        vertical_value = first_value(row, STARTUP_FIELD_HEADERS["vertical_text"])
        if vertical_value:
            self._set_field(org, "vertical_text", str(vertical_value))
            for tag_payload in self.service._split_tags(vertical_value, "VERTICAL"):
                self._ensure_tag_link(db, org, tag_payload["label"], "VERTICAL")

        for field, headers in (
            ("geography_text", STARTUP_FIELD_HEADERS["geography_text"]),
            ("source_text", STARTUP_FIELD_HEADERS["source_text"]),
            ("added_by_text", STARTUP_FIELD_HEADERS["added_by_text"]),
            ("solution_summary", STARTUP_FIELD_HEADERS["solution_summary"]),
        ):
            value = first_value(row, headers)
            if value is not None:
                self._set_field(org, field, str(value))

        last_contact = first_value(row, STARTUP_FIELD_HEADERS["last_contact_date"])
        if isinstance(last_contact, date):
            self._set_field(org, "last_contact_date", last_contact)

        status_value = first_value(row, STARTUP_FIELD_HEADERS["lifecycle_status"])
        status_code = self.service._status_code(status_value, "company_status")
        if status_code:
            status_id = self.service._status_id(db, status_code, "company_status")
            if status_id and org.lifecycle_status_id != status_id:
                if org.lifecycle_status_id is not None and org.updated_by_user_id is not None:
                    self._record_conflict(org, "lifecycle_status", str(org.lifecycle_status_id), status_code)
                else:
                    self._record_change(org, "lifecycle_status", str(org.lifecycle_status_id) if org.lifecycle_status_id else None, status_code)
                    if self.apply_changes:
                        org.lifecycle_status_id = status_id

        for column_name, code in fit_columns.items():
            raw_value = row.get(column_name)
            if raw_value is not None and self.service._is_truthy(raw_value):
                self._ensure_fit(db, org, company_by_code, code, str(raw_value))

        if self.apply_changes:
            db.add(org)

    def _repair_network(self, db, org: Organization, row: dict[str, Any]) -> None:
        subtype = first_value(row, NETWORK_FIELD_HEADERS["organization_subtype"])
        if subtype:
            self._set_field(org, "organization_subtype", str(subtype))
        geography = first_value(row, NETWORK_FIELD_HEADERS["geography_text"])
        if geography:
            self._set_field(org, "geography_text", str(geography))
        added_by = first_value(row, NETWORK_FIELD_HEADERS["added_by_text"])
        if added_by:
            self._set_field(org, "added_by_text", str(added_by))

        relationship_value = first_value(row, NETWORK_FIELD_HEADERS["relationship_status"])
        relationship_code = self.service._status_code(relationship_value, "network_relationship")
        if relationship_code:
            status_id = self.service._status_id(db, relationship_code, "network_relationship")
            if status_id and org.relationship_status_id != status_id:
                if org.relationship_status_id is not None and org.updated_by_user_id is not None:
                    self._record_conflict(org, "relationship_status", str(org.relationship_status_id), relationship_code)
                else:
                    self._record_change(org, "relationship_status", str(org.relationship_status_id) if org.relationship_status_id else None, relationship_code)
                    if self.apply_changes:
                        org.relationship_status_id = status_id

        expertise = first_value(row, NETWORK_FIELD_HEADERS["expertise"])
        for tag_payload in self.service._split_tags(expertise, "EXPERTISE"):
            self._ensure_tag_link(db, org, tag_payload["label"], "EXPERTISE")

        if self.apply_changes:
            db.add(org)

    def _report(self) -> dict[str, Any]:
        return {
            "mode": "APPLY" if self.apply_changes else "DRY_RUN",
            "workbook": str(self.workbook_path),
            "generated_at": datetime.now(timezone.utc).isoformat(),
            "summary": {
                "field_changes": len(self.changes),
                "organizations_touched": len(self.orgs_touched),
                "fit_links_added": self.fit_additions,
                "tag_links_added": self.tag_links_added,
                "conflicts_skipped_manual_edits": len(self.conflicts),
                "workbook_rows_without_db_match": len(self.unmatched),
                "ambiguous_matches_skipped": len(self.ambiguous),
                "workbook_data_quality_issues": len(self.data_quality),
            },
            "changes": self.changes,
            "conflicts": self.conflicts,
            "unmatched_workbook_rows": self.unmatched,
            "ambiguous": self.ambiguous,
            "workbook_data_quality_issues": self.data_quality,
        }


def main() -> None:
    parser = argparse.ArgumentParser(description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter)
    parser.add_argument("workbook", type=Path, help="Path to the source .xlsx workbook")
    parser.add_argument("--apply", action="store_true", help="Write changes to the database (default: dry run)")
    parser.add_argument(
        "--report",
        type=Path,
        default=None,
        help="Where to write the JSON report (default: repair_report_<mode>.json next to this script)",
    )
    args = parser.parse_args()

    if not args.workbook.is_file():
        raise SystemExit(f"Workbook not found: {args.workbook}")

    repair = LibraryRepair(args.workbook, apply_changes=args.apply)
    report = repair.repair()

    report_path = args.report or (
        Path(__file__).resolve().parent / f"repair_report_{'apply' if args.apply else 'dry_run'}.json"
    )
    report_path.write_text(json.dumps(report, ensure_ascii=False, indent=2), encoding="utf-8")

    print(f"Mode: {report['mode']}")
    for key, value in report["summary"].items():
        print(f"  {key}: {value}")
    print(f"Report written to {report_path}")


if __name__ == "__main__":
    main()
