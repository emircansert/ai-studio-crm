import hashlib
import json
import re
import unicodedata
from collections import Counter, defaultdict
from dataclasses import dataclass
from datetime import date, datetime
from enum import StrEnum
from pathlib import Path
from typing import Any
from urllib.parse import urlparse
from uuid import UUID

import yaml
from openpyxl import load_workbook
from sqlalchemy import or_, select
from sqlalchemy.orm import Session

from app.models import ImportBatch, ImportRow, ImportSheet, ImportWarning


class ImportStage(StrEnum):
    UPLOADED = "UPLOADED"
    PROFILED = "PROFILED"
    MAPPED = "MAPPED"
    PREVIEW_READY = "PREVIEW_READY"
    COMMITTED = "COMMITTED"
    REJECTED = "REJECTED"
    FAILED = "FAILED"


@dataclass(frozen=True)
class WorkbookProfile:
    filename: str
    sheets: list[dict[str, Any]]


BLANK_PLACEHOLDERS = {"", "-", "?", "n/a", "na", "none", "null"}
MAX_SAMPLE_ROWS = 5


class ExcelImportPipeline:
    """Workbook profiling, row staging, and preview generation.

    Phase 1 deliberately stops at import staging. It never mutates normalized CRM
    domain tables such as organizations, opportunities, events, or contacts.
    """

    def __init__(self, config_dir: Path) -> None:
        self.config_dir = config_dir
        self.sheet_mapping = self._load_yaml("sheet_mapping.yml")
        self.column_mapping = self._load_yaml("column_mapping.yml")
        self.status_mapping = self._load_yaml("status_mapping.yml")

    def _load_yaml(self, filename: str) -> dict[str, Any]:
        path = self.config_dir / filename
        with path.open("r", encoding="utf-8") as handle:
            return yaml.safe_load(handle) or {}

    def profile_workbook_file(self, workbook_path: Path) -> WorkbookProfile:
        workbook = load_workbook(workbook_path, read_only=True, data_only=True)
        sheets: list[dict[str, Any]] = []
        try:
            for worksheet in workbook.worksheets:
                rows = list(worksheet.iter_rows(values_only=True))
                non_empty_rows = [row for row in rows if self._is_non_empty_row(row)]
                mapping = self._match_sheet(worksheet.title)
                header_row = mapping.get("header_row") if mapping else self._detect_header_row(rows)
                headers = self._headers_from_rows(rows, header_row) if header_row else []
                sheets.append(
                    {
                        "sheet_name": worksheet.title,
                        "detected_entity": mapping.get("entity") if mapping else "UNKNOWN",
                        "header_row": header_row,
                        "row_count": max(len(non_empty_rows) - (1 if headers else 0), 0),
                        "non_empty_row_count": len(non_empty_rows),
                        "headers": headers,
                    }
                )
        finally:
            workbook.close()
        return WorkbookProfile(filename=workbook_path.name, sheets=sheets)

    def stage_workbook(
        self,
        db: Session,
        *,
        workbook_path: Path,
        original_filename: str,
        file_sha256: str,
        file_size_bytes: int,
        uploaded_by_user_id: UUID | None,
    ) -> ImportBatch:
        profile = self.profile_workbook_file(workbook_path)
        batch = ImportBatch(
            original_filename=original_filename,
            file_sha256=file_sha256,
            uploaded_by_user_id=uploaded_by_user_id,
            status=ImportStage.PROFILED.value,
            workbook_metadata={
                "storage_path": str(workbook_path),
                "file_size_bytes": file_size_bytes,
                "detected_sheets": profile.sheets,
                "status_mappings_used": self._status_mapping_summary(),
                "duplicate_candidates": {},
            },
        )
        db.add(batch)
        db.flush()
        prior_file_count = db.execute(
            select(ImportBatch).where(
                ImportBatch.file_sha256 == file_sha256,
                ImportBatch.id != batch.id,
            )
        ).scalars().all()
        if prior_file_count:
            self._add_warning(
                db,
                batch,
                None,
                "WARNING",
                "DUPLICATE_FILE_HASH",
                "A workbook with the same SHA-256 hash has already been uploaded.",
                "file_sha256",
                file_sha256,
            )

        workbook = load_workbook(workbook_path, read_only=True, data_only=True)
        duplicate_trackers: dict[str, dict[str, list[ImportRow]]] = defaultdict(lambda: defaultdict(list))
        duplicate_metadata: dict[str, dict[str, list[dict[str, Any]]]] = defaultdict(lambda: defaultdict(list))
        seen_entities = set()

        try:
            for worksheet in workbook.worksheets:
                mapping = self._match_sheet(worksheet.title)
                entity = mapping.get("entity") if mapping else "UNKNOWN"
                seen_entities.add(entity)
                rows = list(worksheet.iter_rows(values_only=True))
                header_row = mapping.get("header_row") if mapping else self._detect_header_row(rows)
                headers = self._headers_from_rows(rows, header_row) if header_row else []
                non_empty_count = sum(1 for row in rows if self._is_non_empty_row(row))
                import_sheet = ImportSheet(
                    import_batch_id=batch.id,
                    sheet_name=worksheet.title,
                    detected_entity=entity,
                    header_row=header_row,
                    row_count=max(non_empty_count - (1 if headers else 0), 0),
                    column_mapping=self._sheet_column_mapping(entity, headers),
                )
                db.add(import_sheet)
                db.flush()

                if mapping is None:
                    self._add_warning(
                        db,
                        batch,
                        None,
                        "WARNING",
                        "UNKNOWN_SHEET",
                        f"Sheet '{worksheet.title}' is not defined in sheet_mapping.yml.",
                        "sheet_name",
                        worksheet.title,
                    )
                    continue

                self._validate_required_columns(db, batch, worksheet.title, mapping, headers)
                self._warn_unmapped_columns(db, batch, worksheet.title, entity, headers)
                self._validate_column_types(db, batch, worksheet.title, entity, mapping, headers, rows)
                self._stage_sheet_rows(db, batch, import_sheet, mapping, headers, rows, duplicate_trackers)

            self._validate_required_sheets(db, batch, seen_entities)
            self._emit_duplicate_warnings(db, batch, duplicate_trackers, duplicate_metadata)
            batch.status = ImportStage.PREVIEW_READY.value
            batch.workbook_metadata = {
                **(batch.workbook_metadata or {}),
                "duplicate_candidates": {
                    tracker_name: dict(candidates)
                    for tracker_name, candidates in duplicate_metadata.items()
                },
            }
            db.add(batch)
            db.commit()
            db.refresh(batch)
            return batch
        except Exception:
            db.rollback()
            raise
        finally:
            workbook.close()

    def build_preview(self, db: Session, import_batch_id: UUID) -> dict[str, Any]:
        batch = db.get(ImportBatch, import_batch_id)
        if batch is None:
            raise ValueError("Import batch not found")

        sheets = db.execute(
            select(ImportSheet).where(ImportSheet.import_batch_id == batch.id).order_by(ImportSheet.sheet_name)
        ).scalars().all()
        rows_by_sheet: dict[UUID, list[ImportRow]] = {}
        for sheet in sheets:
            rows_by_sheet[sheet.id] = db.execute(
                select(ImportRow).where(ImportRow.import_sheet_id == sheet.id).order_by(ImportRow.excel_row_number)
            ).scalars().all()

        row_ids = [row.id for rows in rows_by_sheet.values() for row in rows]
        if row_ids:
            warning_stmt = select(ImportWarning).where(
                or_(ImportWarning.import_batch_id == batch.id, ImportWarning.import_row_id.in_(row_ids))
            )
        else:
            warning_stmt = select(ImportWarning).where(ImportWarning.import_batch_id == batch.id)
        warnings = db.execute(warning_stmt).scalars().all()

        severity_counts = Counter(w.severity for w in warnings)
        code_counts = Counter(w.code for w in warnings)

        sample_rows: dict[str, list[dict[str, Any]]] = defaultdict(list)
        staged_counts: dict[str, int] = {}
        row_counts: dict[str, int | None] = {}
        detected_sheets: list[dict[str, Any]] = []
        for sheet in sheets:
            entity = sheet.detected_entity or "UNKNOWN"
            sheet_rows = rows_by_sheet[sheet.id]
            row_counts[sheet.sheet_name] = sheet.row_count
            staged_counts[entity] = staged_counts.get(entity, 0) + len(sheet_rows)
            detected_sheets.append(
                {
                    "id": str(sheet.id),
                    "sheet_name": sheet.sheet_name,
                    "detected_entity": entity,
                    "header_row": sheet.header_row,
                    "row_count": sheet.row_count,
                    "staged_row_count": len(sheet_rows),
                    "column_mapping": sheet.column_mapping,
                }
            )
            if len(sample_rows[entity]) < MAX_SAMPLE_ROWS:
                for row in sheet_rows[: MAX_SAMPLE_ROWS - len(sample_rows[entity])]:
                    sample_rows[entity].append(
                        {
                            "id": str(row.id),
                            "excel_row_number": row.excel_row_number,
                            "validation_status": row.validation_status,
                            "cleaned_values": row.cleaned_values,
                            "normalized_candidate": row.normalized_candidate,
                        }
                    )

        missing_mappings = [
            {"code": w.code, "message": w.message, "field_name": w.field_name, "raw_value": w.raw_value}
            for w in warnings
            if w.code
            in {
                "UNKNOWN_SHEET",
                "MISSING_REQUIRED_SHEET",
                "MISSING_REQUIRED_COLUMN",
                "UNMAPPED_COLUMN",
                "COLUMN_TYPE_MISMATCH",
            }
        ]

        return {
            "batch": {
                "id": str(batch.id),
                "original_filename": batch.original_filename,
                "file_sha256": batch.file_sha256,
                "status": batch.status,
                "created_at": batch.created_at.isoformat() if batch.created_at else None,
                "updated_at": batch.updated_at.isoformat() if batch.updated_at else None,
                "workbook_metadata": batch.workbook_metadata,
            },
            "detected_sheets": detected_sheets,
            "row_counts_by_sheet": row_counts,
            "staged_row_counts": staged_counts,
            "warning_counts": {
                "by_severity": dict(severity_counts),
                "by_code": dict(code_counts),
            },
            "sample_rows": dict(sample_rows),
            "duplicate_candidates": (batch.workbook_metadata or {}).get("duplicate_candidates", {}),
            "missing_mappings_or_columns": missing_mappings,
            "status_mappings_used": (batch.workbook_metadata or {}).get("status_mappings_used", {}),
        }

    def commit(self, import_batch_id: UUID, confirmed_by_user_id: UUID) -> None:
        raise NotImplementedError("Import commit is intentionally deferred until Excel Import Phase 2.")

    def _match_sheet(self, sheet_name: str) -> dict[str, Any] | None:
        normalized_sheet_name = self._normalize_key(sheet_name)
        for config in (self.sheet_mapping.get("sheets") or {}).values():
            source_names = config.get("source_names") or []
            if normalized_sheet_name in {self._normalize_key(name) for name in source_names}:
                return config
        return None

    def _detect_header_row(self, rows: list[tuple[Any, ...]]) -> int | None:
        best_row = None
        best_count = 0
        for index, row in enumerate(rows[:20], start=1):
            count = sum(1 for value in row if self._clean_cell(value) is not None)
            if count > best_count:
                best_count = count
                best_row = index
        return best_row if best_count else None

    def _headers_from_rows(self, rows: list[tuple[Any, ...]], header_row: int | None) -> list[str]:
        if not header_row or header_row > len(rows):
            return []
        headers: list[str] = []
        for index, value in enumerate(rows[header_row - 1], start=1):
            cleaned = self._clean_cell(value)
            headers.append(str(cleaned) if cleaned is not None else f"__EMPTY_{index}")
        while headers and headers[-1].startswith("__EMPTY_"):
            headers.pop()
        return headers

    def _sheet_column_mapping(self, entity: str, headers: list[str]) -> dict[str, Any]:
        configured = ((self.column_mapping.get("entities") or {}).get(entity) or {}).get("columns") or {}
        mapped = {
            header: configured.get(header, {"target": None, "transform": None})
            for header in headers
            if not header.startswith("__EMPTY_")
        }
        missing_config = [header for header in headers if header not in configured and not header.startswith("__EMPTY_")]
        return {"mapped_columns": mapped, "missing_config_columns": missing_config}

    def _stage_sheet_rows(
        self,
        db: Session,
        batch: ImportBatch,
        import_sheet: ImportSheet,
        mapping: dict[str, Any],
        headers: list[str],
        rows: list[tuple[Any, ...]],
        duplicate_trackers: dict[str, dict[str, list[ImportRow]]],
    ) -> None:
        first_data_row = int(mapping.get("first_data_row") or (int(mapping.get("header_row") or 1) + 1))
        entity = mapping["entity"]
        for excel_row_number, row in enumerate(rows, start=1):
            if excel_row_number < first_data_row:
                continue
            if not self._is_non_empty_row(row):
                self._add_warning(
                    db,
                    batch,
                    None,
                    "INFO",
                    "EMPTY_ROW_SKIPPED",
                    f"Empty row {excel_row_number} was skipped in sheet '{import_sheet.sheet_name}'.",
                )
                continue
            if self._matches_skip_rule(row, headers, mapping):
                self._add_warning(
                    db,
                    batch,
                    None,
                    "INFO",
                    "EMPTY_ROW_SKIPPED",
                    f"Non-record row {excel_row_number} was skipped in sheet '{import_sheet.sheet_name}'.",
                )
                continue

            raw_values = self._row_to_dict(row, headers, preserve_raw=True)
            cleaned_values = self._row_to_dict(row, headers, preserve_raw=False)
            normalized_candidate = self._build_normalized_candidate(entity, cleaned_values)
            import_row = ImportRow(
                import_sheet_id=import_sheet.id,
                excel_row_number=excel_row_number,
                raw_values=raw_values,
                cleaned_values=cleaned_values,
                normalized_candidate=normalized_candidate,
                row_hash=self._hash_dict(raw_values),
                validation_status="VALID",
            )
            db.add(import_row)
            db.flush()

            warning_count = self._validate_row(db, batch, import_row, entity, cleaned_values)
            if warning_count:
                import_row.validation_status = "WARNING"

            normalized_name = normalized_candidate.get("normalized_name")
            if entity in {"STARTUP_LIBRARY", "NETWORK"} and normalized_name:
                duplicate_trackers["normalized_name"][normalized_name].append(import_row)
            website_domain = normalized_candidate.get("website_domain")
            if entity == "STARTUP_LIBRARY" and website_domain:
                duplicate_trackers["website_domain"][website_domain].append(import_row)

    def _validate_required_columns(
        self,
        db: Session,
        batch: ImportBatch,
        sheet_name: str,
        mapping: dict[str, Any],
        headers: list[str],
    ) -> None:
        header_set = set(headers)
        for requirement in mapping.get("required_columns") or []:
            # A requirement may be a single header name or a list of alternative
            # names (older Turkish headers vs newer English headers).
            alternatives = requirement if isinstance(requirement, list) else [requirement]
            if not any(alternative in header_set for alternative in alternatives):
                display_name = " / ".join(str(alternative) for alternative in alternatives)
                self._add_warning(
                    db,
                    batch,
                    None,
                    "ERROR",
                    "MISSING_REQUIRED_COLUMN",
                    f"Sheet '{sheet_name}' is missing required column '{display_name}'.",
                    "column_name",
                    display_name,
                )

    def _warn_unmapped_columns(
        self,
        db: Session,
        batch: ImportBatch,
        sheet_name: str,
        entity: str,
        headers: list[str],
    ) -> None:
        """Warn when a known sheet carries headers the mapping config does not recognise.

        Renamed or shifted columns are the main way a workbook silently loses data on
        import, so surface them clearly during preview instead of ignoring them.
        """
        configured = ((self.column_mapping.get("entities") or {}).get(entity) or {}).get("columns") or {}
        if not configured:
            return
        unmapped = [
            header
            for header in headers
            if header not in configured and not header.startswith("__EMPTY_")
        ]
        if unmapped:
            self._add_warning(
                db,
                batch,
                None,
                "WARNING",
                "UNMAPPED_COLUMN",
                (
                    f"Sheet '{sheet_name}' has {len(unmapped)} column(s) that are not in the import mapping "
                    f"and will be ignored: {', '.join(unmapped)}. "
                    "If these should be imported, the column mapping needs updating."
                ),
                "columns",
                ", ".join(unmapped)[:500],
            )

    def _validate_column_types(
        self,
        db: Session,
        batch: ImportBatch,
        sheet_name: str,
        entity: str,
        mapping: dict[str, Any],
        headers: list[str],
        rows: list[tuple[Any, ...]],
    ) -> None:
        """Detect probable column misalignment by sanity-checking date columns.

        If a column mapped with a strict date transform holds mostly non-date values,
        the workbook layout almost certainly shifted; block-level ERROR makes this
        loud instead of silently importing wrong values.
        """
        configured = ((self.column_mapping.get("entities") or {}).get(entity) or {}).get("columns") or {}
        date_transforms = {"parse_date", "parse_date_with_warning"}
        first_data_row = int(mapping.get("first_data_row") or (int(mapping.get("header_row") or 1) + 1))
        for index, header in enumerate(headers):
            column_config = configured.get(header) or {}
            if column_config.get("transform") not in date_transforms:
                continue
            non_empty = 0
            date_like = 0
            for row_number, row in enumerate(rows, start=1):
                if row_number < first_data_row or index >= len(row):
                    continue
                value = self._clean_cell(row[index])
                if value is None:
                    continue
                non_empty += 1
                if isinstance(value, str) and self._looks_like_iso_date(value):
                    date_like += 1
                elif isinstance(value, (datetime, date)):
                    date_like += 1
            if non_empty >= 5 and date_like / non_empty < 0.5:
                self._add_warning(
                    db,
                    batch,
                    None,
                    "ERROR",
                    "COLUMN_TYPE_MISMATCH",
                    (
                        f"Sheet '{sheet_name}' column '{header}' should contain dates but "
                        f"{non_empty - date_like} of {non_empty} values are not dates. "
                        "The workbook columns may be misaligned; review before committing."
                    ),
                    "column_name",
                    header,
                )

    def _validate_required_sheets(self, db: Session, batch: ImportBatch, seen_entities: set[str]) -> None:
        for config in (self.sheet_mapping.get("sheets") or {}).values():
            entity = config.get("entity")
            if config.get("allow_empty") is True:
                continue
            if entity and entity not in seen_entities:
                self._add_warning(
                    db,
                    batch,
                    None,
                    "ERROR",
                    "MISSING_REQUIRED_SHEET",
                    f"Required sheet for entity '{entity}' was not found.",
                    "entity",
                    entity,
                )

    def _validate_row(
        self,
        db: Session,
        batch: ImportBatch,
        import_row: ImportRow,
        entity: str,
        values: dict[str, Any],
    ) -> int:
        warning_count = 0

        if entity == "STARTUP_LIBRARY":
            if not values.get("Website"):
                warning_count += self._add_warning(
                    db, batch, import_row, "WARNING", "MISSING_WEBSITE", "Startup row has no website.", "Website"
                )
            else:
                domain = self._extract_domain(values.get("Website"))
                if domain is None:
                    warning_count += self._add_warning(
                        db,
                        batch,
                        import_row,
                        "WARNING",
                        "MALFORMED_WEBSITE_URL",
                        "Website value could not be parsed into a valid domain.",
                        "Website",
                        str(values.get("Website")),
                    )
            startup_contact_column = self._first_present_column(values, "Contact Person", "Kontak Kişisi")
            startup_contact = values.get(startup_contact_column)
            if startup_contact and not self._extract_email(str(startup_contact)):
                warning_count += self._add_warning(
                    db,
                    batch,
                    import_row,
                    "WARNING",
                    "CONTACT_WITHOUT_EMAIL",
                    "Contact text does not contain an email address.",
                    startup_contact_column,
                    str(startup_contact),
                )
            warning_count += self._validate_status_value(
                db,
                batch,
                import_row,
                values,
                self._first_present_column(values, "Process Status (1 to 7)", "Status"),
                "company_status",
            )
            warning_count += self._validate_date_value(db, batch, import_row, values, "Last Contact")

        elif entity == "POC_OPPORTUNITIES":
            warning_count += self._validate_status_value(
                db,
                batch,
                import_row,
                values,
                self._first_present_column(values, "Status", "Son Durum"),
                "opportunity_stage",
            )
            warning_count += self._validate_date_value(
                db,
                batch,
                import_row,
                values,
                self._first_present_column(values, "Last Contact", "Son Görüşme"),
            )

        elif entity == "EVENTS":
            warning_count += self._validate_status_value(db, batch, import_row, values, "AI Program Relevance", "ratings")
            warning_count += self._validate_status_value(
                db,
                batch,
                import_row,
                values,
                self._first_present_column(values, "Value Creation Potential", "Değer yaratma Opsiyonu"),
                "ratings",
            )
            warning_count += self._validate_date_value(db, batch, import_row, values, "Date")

        elif entity == "NETWORK":
            warning_count += self._validate_status_value(db, batch, import_row, values, "Relationship", "network_relationship")
            if values.get("Contact Person") and not self._extract_email(str(values.get("Contact Person"))):
                warning_count += self._add_warning(
                    db,
                    batch,
                    import_row,
                    "WARNING",
                    "CONTACT_WITHOUT_EMAIL",
                    "Contact text does not contain an email address.",
                    "Contact Person",
                    str(values.get("Contact Person")),
                )

        return warning_count

    def _validate_status_value(
        self,
        db: Session,
        batch: ImportBatch,
        import_row: ImportRow,
        values: dict[str, Any],
        column_name: str,
        vocabulary: str,
    ) -> int:
        raw_value = values.get(column_name)
        if raw_value is None:
            return 0
        alias_map = self._status_alias_map(vocabulary)
        if self._normalize_key(str(raw_value)) not in alias_map:
            return self._add_warning(
                db,
                batch,
                import_row,
                "ERROR",
                "UNKNOWN_STATUS",
                f"Unknown status value '{raw_value}' for vocabulary '{vocabulary}'.",
                column_name,
                str(raw_value),
            )
        return 0

    def _validate_date_value(
        self,
        db: Session,
        batch: ImportBatch,
        import_row: ImportRow,
        values: dict[str, Any],
        column_name: str,
    ) -> int:
        raw_value = values.get(column_name)
        if raw_value is None:
            return self._add_warning(
                db,
                batch,
                import_row,
                "WARNING",
                "PARTIAL_DATE",
                f"Date field '{column_name}' is blank.",
                column_name,
            )
        if isinstance(raw_value, str) and not self._looks_like_iso_date(raw_value):
            return self._add_warning(
                db,
                batch,
                import_row,
                "WARNING",
                "PARTIAL_DATE",
                f"Date field '{column_name}' is preserved as text and needs review.",
                column_name,
                raw_value,
            )
        return 0

    def _emit_duplicate_warnings(
        self,
        db: Session,
        batch: ImportBatch,
        duplicate_trackers: dict[str, dict[str, list[ImportRow]]],
        duplicate_metadata: dict[str, dict[str, list[dict[str, Any]]]],
    ) -> None:
        duplicate_specs = {
            "normalized_name": ("DUPLICATE_NORMALIZED_COMPANY_NAME_CANDIDATE", "Duplicate normalized company name candidate."),
            "website_domain": ("DUPLICATE_WEBSITE_DOMAIN_CANDIDATE", "Duplicate website domain candidate."),
        }
        for tracker_name, values in duplicate_trackers.items():
            code, message = duplicate_specs[tracker_name]
            for candidate, rows in values.items():
                if len(rows) < 2:
                    continue
                duplicate_metadata[tracker_name][candidate] = [
                    {"import_row_id": str(row.id), "excel_row_number": row.excel_row_number}
                    for row in rows
                ]
                for row in rows:
                    self._add_warning(
                        db,
                        batch,
                        row,
                        "WARNING",
                        code,
                        message,
                        tracker_name,
                        candidate,
                    )
                    row.validation_status = "WARNING"

    def _add_warning(
        self,
        db: Session,
        batch: ImportBatch,
        import_row: ImportRow | None,
        severity: str,
        code: str,
        message: str,
        field_name: str | None = None,
        raw_value: str | None = None,
    ) -> int:
        db.add(
            ImportWarning(
                import_batch_id=batch.id,
                import_row_id=import_row.id if import_row else None,
                severity=severity,
                code=code,
                message=message,
                field_name=field_name,
                raw_value=raw_value,
            )
        )
        return 1

    def _row_to_dict(self, row: tuple[Any, ...], headers: list[str], *, preserve_raw: bool) -> dict[str, Any]:
        values: dict[str, Any] = {}
        for index, header in enumerate(headers):
            if header.startswith("__EMPTY_"):
                continue
            cell_value = row[index] if index < len(row) else None
            cleaned = self._serialize_cell(cell_value) if preserve_raw else self._clean_cell(cell_value)
            # Some sheets repeat a header name (e.g. the PoC sheet has two "Status"
            # columns in side-by-side blocks). Keep the first non-empty value instead
            # of letting a later empty column silently overwrite real data.
            if header in values and values[header] is not None:
                continue
            values[header] = cleaned
        return values

    def _build_normalized_candidate(self, entity: str, values: dict[str, Any]) -> dict[str, Any]:
        name_columns: dict[str, tuple[str, ...]] = {
            "STARTUP_LIBRARY": ("Startup",),
            "POC_OPPORTUNITIES": ("Startup",),
            "NETWORK": ("Institution",),
            "AI_TOOLS": ("Tool Name",),
            "EVENTS": ("Event Name", "Etkinlik Adı"),
        }
        name_value = self._first_value(values, *name_columns.get(entity, ()))
        website = values.get("Website")
        return {
            "entity_type": entity,
            "name": name_value,
            "normalized_name": self._normalize_company_name(str(name_value)) if name_value else None,
            "website_domain": self._extract_domain(str(website)) if website else None,
        }

    def _matches_skip_rule(self, row: tuple[Any, ...], headers: list[str], mapping: dict[str, Any]) -> bool:
        skip_rules = mapping.get("skip_rules") or []
        if not skip_rules:
            return False
        non_empty_indexes = [index for index, value in enumerate(row) if self._clean_cell(value) is not None]
        for rule in skip_rules:
            if rule.get("type") == "non_record_tail_value" and non_empty_indexes == [8]:
                return True
        return False

    def _first_value(self, values: dict[str, Any], *column_names: str) -> Any:
        """Return the first non-empty value among alternative header names."""
        for column_name in column_names:
            value = values.get(column_name)
            if value not in (None, ""):
                return value
        return None

    def _first_present_column(self, values: dict[str, Any], *column_names: str) -> str:
        """Return the first alternative header that carries a value, else the first name."""
        for column_name in column_names:
            if values.get(column_name) not in (None, ""):
                return column_name
        return column_names[0]

    def _is_non_empty_row(self, row: tuple[Any, ...]) -> bool:
        return any(self._clean_cell(value) is not None for value in row)

    def _clean_cell(self, value: Any) -> Any:
        if value is None:
            return None
        if isinstance(value, datetime):
            return value.date().isoformat()
        if isinstance(value, date):
            return value.isoformat()
        if isinstance(value, str):
            cleaned = " ".join(value.replace("\xa0", " ").strip().split())
            if cleaned.lower() in BLANK_PLACEHOLDERS:
                return None
            return cleaned
        return value

    def _serialize_cell(self, value: Any) -> Any:
        if isinstance(value, datetime):
            return value.isoformat()
        if isinstance(value, date):
            return value.isoformat()
        if isinstance(value, str):
            return value
        return value

    def _hash_dict(self, values: dict[str, Any]) -> str:
        payload = json.dumps(values, sort_keys=True, ensure_ascii=False, default=str)
        return hashlib.sha256(payload.encode("utf-8")).hexdigest()

    def _normalize_key(self, value: str) -> str:
        normalized = unicodedata.normalize("NFKC", value or "").casefold()
        return " ".join(normalized.strip().split())

    def _normalize_company_name(self, value: str) -> str:
        normalized = self._normalize_key(value)
        normalized = re.sub(r"[^\w\s.-]", "", normalized, flags=re.UNICODE)
        return " ".join(normalized.split())

    def _extract_domain(self, value: str) -> str | None:
        candidate = value.strip()
        if not candidate:
            return None
        if " " in candidate:
            return None
        if "://" not in candidate:
            candidate = f"https://{candidate}"
        parsed = urlparse(candidate)
        domain = (parsed.netloc or "").split("@")[-1].split(":")[0].lower()
        if domain.startswith("www."):
            domain = domain[4:]
        if "." not in domain or len(domain) > 255:
            return None
        return domain

    def _extract_email(self, value: str) -> str | None:
        match = re.search(r"[A-Z0-9._%+-]+@[A-Z0-9.-]+\.[A-Z]{2,}", value, re.IGNORECASE)
        return match.group(0).lower() if match else None

    def _looks_like_iso_date(self, value: str) -> bool:
        try:
            date.fromisoformat(value)
        except ValueError:
            return False
        return True

    def _status_alias_map(self, vocabulary: str) -> dict[str, str]:
        aliases: dict[str, str] = {}
        for code, config in (self.status_mapping.get(vocabulary) or {}).items():
            aliases[self._normalize_key(code)] = code
            aliases[self._normalize_key(config.get("label", code))] = code
            for alias in config.get("aliases") or []:
                if alias not in (None, ""):
                    aliases[self._normalize_key(str(alias))] = code
        return aliases

    def _status_mapping_summary(self) -> dict[str, Any]:
        summary: dict[str, Any] = {}
        for vocabulary, statuses in self.status_mapping.items():
            if vocabulary == "version" or not isinstance(statuses, dict):
                continue
            summary[vocabulary] = {
                code: {"label": config.get("label"), "alias_count": len(config.get("aliases") or [])}
                for code, config in statuses.items()
            }
        return summary
